import openpyxl as xl

dirList = 'D:\IUT\Semestre5\SAE\database.xlsx'

file = xl.open(dirList)
workSheet = file['Sheet1']

i=2
while (workSheet.cell(i,1).value!= None) :
    name = workSheet.cell(i,1).value
    if (name[0].isalpha()):
        name=name[0].upper() + name[1:]
    workSheet.cell(i,2).value = name
    i=i+1

file.save(dirList)